# Файл: app/handlers.py

from datetime import datetime, time
import io
import dateparser
import openpyxl
from aiogram import F, Router, Bot
from aiogram.filters import CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import (CallbackQuery, Message,
                           ReplyKeyboardRemove,
                           InlineKeyboardMarkup, InlineKeyboardButton)
from aiogram.types.input_file import BufferedInputFile
from aiogram.types import FSInputFile
from apscheduler.schedulers.asyncio import AsyncIOScheduler # Рекомендую убрать или перенести в main.py
from haversine import Unit, haversine
import logging

import app.keyboards as kb
import database.requests as bd
import sites.purse as purse


class Student(StatesGroup):
    education = State()
    institute = State()
    fuclty = State()
    group = State()
    course_for_group_selection = State()
    name = State()
    schedule = State()


class AddData(StatesGroup):
    waiting_for_institute_name = State()
    waiting_for_fuclty_name = State()
    waiting_for_group_name = State()
    waiting_for_group_url = State()
    waiting_for_student_fio = State()


router = Router()
ADMIN_USER_ID = 1033042355

start_text = """
Привет!😁✌️
Здесь ты можешь:
- отметиться в списке присутвующих на занятии
- получить расписание на неделю
- выгрузить список присутствующих за этот учебный год
"""
info_text = """Для навигации используй кнопки\n
📌<b>"Авторизоваться"</b> - чтобы найти себя в базе
📌<b>"Информация"</b> - чтобы вывести инфу о себе
📌<b>"Отметиться"</b> - чтобы отметиться на паре
учти, что отмечаться можно за 20 минут до начала занятия, а также проверяется геолокация😉 
📌<b>"Расписание"</b> - чтобы посмотреть расписание на текущую неделю
📌<b>"Списки присутствующих"</b> - чтобы создался файл со списками присутствующих"""


async def clear_schedule_job_placeholder(bot_instance: Bot):  # Заглушка, если APScheduler вынесен
    logging.info("APScheduler: Плановая задача clear_schedule_job вызвана (заглушка).")
    # Здесь может быть логика, не зависящая от FSMContext конкретного пользователя,
    # например, очистка старых кэшированных данных в БД или глобальном кэше.


@router.message(CommandStart())
async def command_start(message: Message, state: FSMContext) -> None:
    # Логику APScheduler лучше перенести в main.py и настроить корректно
    # scheduler = AsyncIOScheduler(timezone="Europe/Moscow")
    # scheduler.add_job(clear_schedule_job_placeholder, "cron", minute=1, hour=0, day_of_week="mon", args=[message.bot])
    # try:
    #     if not scheduler.running: # Запускаем только если еще не запущен
    #        scheduler.start()
    # except Exception as e_sched:
    #     logging.error(f"Ошибка запуска/работы APScheduler: {e_sched}")
    try:
        await message.delete()
    except Exception:
        pass
    await message.answer(start_text, reply_markup=kb.first_kb)
    await message.answer(info_text, parse_mode="html")


@router.message(F.text == "Авторизоваться")
async def start_authorization_handler(message: Message, state: FSMContext) -> None:
    try:
        await message.delete()
    except Exception:
        pass
    sent_msg = await message.answer("Начинаем процесс авторизации...", reply_markup=ReplyKeyboardRemove())
    try:
        await sent_msg.delete()
    except Exception:
        pass
    await state.clear()
    await state.set_state(Student.education)
    await message.answer("1. Выберите форму обучения:", reply_markup=kb.get_education_kb())


@router.callback_query(
    StateFilter(None, Student.education, Student.institute, Student.fuclty, Student.group, Student.name, AddData),
    F.data == "cancel_auth_process")
async def cancel_authorization_process(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    try:
        await callback.message.edit_text("Авторизация отменена. Вы возвращены в главное меню.")
    except Exception:
        try:  # Если edit_text не сработал, удаляем старое и отправляем новое
            await callback.message.delete()
            await callback.message.answer("Авторизация отменена. Вы возвращены в главное меню.")
        except Exception:
            # Если и удаление не сработало, просто отвечаем на колбэк
            pass
    await callback.message.answer("Выберите действие:", reply_markup=kb.first_kb)
    await callback.answer("Авторизация отменена")


@router.callback_query(Student.education, F.data.startswith("education_"))
async def process_education_choice(callback: CallbackQuery, state: FSMContext):
    education_choice = callback.data.split("education_")[1]
    await state.update_data(education=education_choice)
    await state.set_state(Student.institute)
    await callback.message.edit_text(
        f"Форма обучения: {education_choice.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"2. Выберите ваш институт:",
        reply_markup=kb.get_institute_kb(education_choice, 1)
    )
    await callback.answer()


@router.callback_query(Student.institute, F.data == "back_to_education_selection")
async def back_to_education_from_institute(callback: CallbackQuery, state: FSMContext):
    await state.set_state(Student.education)
    await state.update_data(institute_id=None, education=None)
    await callback.message.edit_text("1. Выберите форму обучения:", reply_markup=kb.get_education_kb())
    await callback.answer()


@router.callback_query(Student.institute, F.data.startswith("institute_page_"))
async def paginate_institutes(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split("_")[2])
    data = await state.get_data()
    education_state = data.get("education")
    if not education_state:
        await state.set_state(Student.education)
        await callback.message.edit_text("Ошибка: форма обучения не найдена. Начните сначала.",
                                         reply_markup=kb.get_education_kb())
        await callback.answer("Ошибка состояния", show_alert=True)
        return
    await callback.message.edit_text(
        f"Форма обучения: {education_state.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"2. Выберите ваш институт (стр. {page}):",
        reply_markup=kb.get_institute_kb(education_state, page)
    )
    await callback.answer()


@router.callback_query(Student.institute, F.data.startswith("id_institute_"))
async def process_institute_choice(callback: CallbackQuery, state: FSMContext):
    institute_id = int(callback.data.split("_")[-1])
    await state.update_data(institute_id=institute_id)
    data = await state.get_data()
    education_state = data.get("education")
    institute_name = bd.get_institute_from_id(institute_id)
    await state.set_state(Student.fuclty)
    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"Институт: {institute_name}\n"
        f"3. Выберите ваше направление:",
        reply_markup=kb.get_fuclty_kb(institute_id, education_state, 1)
    )
    await callback.answer()


@router.callback_query(Student.fuclty, F.data.startswith("back_to_institute_selection_"))
async def back_to_institute_from_faculty(callback: CallbackQuery, state: FSMContext):
    education_state = callback.data.split("_")[-1]
    await state.set_state(Student.institute)
    # institute_id сбрасывается при переходе в Student.institute, так что его не нужно передавать
    # education уже должен быть в state или мы его устанавливаем из callback_data
    await state.update_data(fuclty_id=None, education=education_state) # Убедимся, что education в state
    await callback.message.edit_text(
        f"Форма обучения: {education_state.replace('бакалавр','Бакалавриат').replace('магистр','Магистратура')}\n"
        f"2. Выберите ваш институт:", # institute_name здесь не нужен, т.к. мы выбираем институт
        reply_markup=kb.get_institute_kb(education_state, 1)
    )
    await callback.answer()


@router.callback_query(Student.fuclty, F.data.startswith("fuclty_page_"))
async def paginate_faculties(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split("_")[2])
    data = await state.get_data()
    institute_id_state = data.get("institute_id")
    education_state = data.get("education")
    if not institute_id_state or not education_state:
        await state.set_state(Student.education)
        await callback.message.edit_text("Ошибка: Начните сначала.", reply_markup=kb.get_education_kb())
        await callback.answer("Ошибка состояния", show_alert=True)
        return
    institute_name = bd.get_institute_from_id(institute_id_state)
    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"Институт: {institute_name}\n"
        f"3. Выберите ваше направление (стр. {page}):",
        reply_markup=kb.get_fuclty_kb(institute_id_state, education_state, page)
    )
    await callback.answer()


@router.callback_query(Student.fuclty, F.data.startswith("id_fuclty_"))
async def process_faculty_choice(callback: CallbackQuery, state: FSMContext):
    faculty_id = int(callback.data.split("_")[-1])
    await state.update_data(fuclty_id=faculty_id, course_for_group_selection=1)
    data = await state.get_data()
    education_state = data.get("education")
    institute_id_state = data.get("institute_id")
    faculty_name = bd.get_fuclty_from_id(faculty_id)
    institute_name = bd.get_institute_from_id(institute_id_state)
    await state.set_state(Student.group)
    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"Институт: {institute_name}\n"
        f"Направление: {faculty_name}\n"
        f"4. Выберите вашу группу (1 курс):",
        reply_markup=kb.get_group_kb(faculty_id, 1, education_state, 1)
    )
    await callback.answer()


@router.callback_query(Student.group, F.data == "back_to_faculty_selection")
async def back_to_faculty_from_group(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    institute_id_state = data.get("institute_id")
    education_state = data.get("education")
    if not institute_id_state or not education_state:
        await state.set_state(Student.education)
        await callback.message.edit_text("Ошибка: Начните сначала.", reply_markup=kb.get_education_kb())
        await callback.answer("Ошибка состояния", show_alert=True)
        return
    await state.set_state(Student.fuclty)
    await state.update_data(study_group_id=None, course_for_group_selection=None)
    institute_name = bd.get_institute_from_id(institute_id_state) # institute_name определяется здесь
    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр','Бакалавриат').replace('магистр','Магистратура')}\n"
        f"Институт: {institute_name}\n" # Используем определенный institute_name
        f"3. Выберите ваше направление:",
        reply_markup=kb.get_fuclty_kb(institute_id_state, education_state, 1)
    )
    await callback.answer()


@router.callback_query(Student.group, F.data.startswith("course_"))
async def process_group_course_choice(callback: CallbackQuery, state: FSMContext):
    course = int(callback.data.split("_")[1])
    await state.update_data(course_for_group_selection=course)
    data = await state.get_data()
    faculty_id_state = data.get("fuclty_id")
    education_state = data.get("education")
    institute_id_state = data.get("institute_id")
    if not all([faculty_id_state, education_state, institute_id_state]):
        await state.set_state(Student.education)
        await callback.message.edit_text("Ошибка: Начните сначала.", reply_markup=kb.get_education_kb())
        await callback.answer("Ошибка состояния", show_alert=True)
        return
    institute_name = bd.get_institute_from_id(institute_id_state)
    faculty_name = bd.get_fuclty_from_id(faculty_id_state)
    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"Институт: {institute_name}\n"
        f"Направление: {faculty_name}\n"
        f"4. Выберите вашу группу ({course} курс):",
        reply_markup=kb.get_group_kb(faculty_id_state, course, education_state, 1)
    )
    await callback.answer()


@router.callback_query(Student.group, F.data.startswith("study_group_page_"))
async def paginate_groups(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    page = int(parts[3])
    course = int(parts[4])
    data = await state.get_data()
    faculty_id_state = data.get("fuclty_id")
    education_state = data.get("education")
    institute_id_state = data.get("institute_id")
    if not all([faculty_id_state, education_state, institute_id_state]):
        await state.set_state(Student.education)
        await callback.message.edit_text("Ошибка: Начните сначала.", reply_markup=kb.get_education_kb())
        await callback.answer("Ошибка состояния", show_alert=True)
        return
    institute_name = bd.get_institute_from_id(institute_id_state)
    faculty_name = bd.get_fuclty_from_id(faculty_id_state)
    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"Институт: {institute_name}\n"
        f"Направление: {faculty_name}\n"
        f"4. Выберите вашу группу ({course} курс, стр. {page}):",
        reply_markup=kb.get_group_kb(faculty_id_state, course, education_state, page)
    )
    await callback.answer()


@router.callback_query(Student.group, F.data.startswith("id_study_group_"))
async def process_group_choice(callback: CallbackQuery, state: FSMContext):
    study_group_id = int(callback.data.split("_")[-1])  # Исправлен индекс
    await state.update_data(study_group_id=study_group_id)
    data = await state.get_data()
    education_state = data.get("education")
    institute_id_state = data.get("institute_id")
    faculty_id_state = data.get("fuclty_id")
    institute_name = bd.get_institute_from_id(institute_id_state)
    faculty_name = bd.get_fuclty_from_id(faculty_id_state)
    group_name = bd.get_study_group_from_id(study_group_id)
    await state.set_state(Student.name)
    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"Институт: {institute_name}\n"
        f"Направление: {faculty_name}\n"
        f"Группа: {group_name}\n"
        f"5. Выберите себя из списка:",
        reply_markup=kb.get_student_kb(study_group_id, 1)
    )
    await callback.answer()


@router.callback_query(Student.name, F.data == "back_to_group_selection")
async def back_to_group_from_student(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    faculty_id_state = data.get("fuclty_id")
    education_state = data.get("education")
    institute_id_state = data.get("institute_id")
    course = data.get("course_for_group_selection", 1)
    if not all([faculty_id_state, education_state, institute_id_state]):
        await state.set_state(Student.education)
        await callback.message.edit_text("Ошибка: Начните сначала.", reply_markup=kb.get_education_kb())
        await callback.answer("Ошибка состояния", show_alert=True)
        return
    await state.set_state(Student.group)
    await state.update_data(student_id=None)
    institute_name = bd.get_institute_from_id(institute_id_state) # institute_name определяется здесь
    faculty_name = bd.get_fuclty_from_id(faculty_id_state)     # faculty_name определяется здесь
    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр','Бакалавриат').replace('магистр','Магистратура')}\n"
        f"Институт: {institute_name}\n" # Используем
        f"Направление: {faculty_name}\n" # Используем
        f"4. Выберите вашу группу ({course} курс):",
        reply_markup=kb.get_group_kb(faculty_id_state, course, education_state, 1)
    )
    await callback.answer()


@router.callback_query(Student.name, F.data.startswith("student_page_"))
async def paginate_students(callback: CallbackQuery, state: FSMContext):
    page = int(callback.data.split("_")[2])
    data = await state.get_data()
    study_group_id_state = data.get("study_group_id")
    if not study_group_id_state:
        await state.set_state(Student.group)
        # Здесь нужно корректно восстановить предыдущий шаг или отправить в начало
        await callback.message.edit_text("Ошибка: Группа не выбрана. Пожалуйста, выберите группу.", reply_markup=None)
        await callback.answer("Ошибка состояния", show_alert=True)
        return

    education_state = data.get("education")
    institute_id_state = data.get("institute_id")
    faculty_id_state = data.get("fuclty_id")
    institute_name = bd.get_institute_from_id(institute_id_state) if institute_id_state else ""
    faculty_name = bd.get_fuclty_from_id(faculty_id_state) if faculty_id_state else ""
    group_name = bd.get_study_group_from_id(study_group_id_state) if study_group_id_state else ""

    await callback.message.edit_text(
        f"Форма: {education_state.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура') if education_state else ''}\n"
        f"Институт: {institute_name}\n"
        f"Направление: {faculty_name}\n"
        f"Группа: {group_name}\n"
        f"5. Выберите себя из списка (стр. {page}):",
        reply_markup=kb.get_student_kb(study_group_id_state, page)
    )
    await callback.answer()


@router.callback_query(Student.name, F.data.startswith("id_student_"))
async def process_student_choice(callback: CallbackQuery, state: FSMContext):
    student_id = int(callback.data.split("_")[-1])  # Исправлен индекс
    await state.update_data(student_id=student_id)
    await state.set_state(None)  # Завершение FSM авторизации

    try:
        await callback.message.delete()
    except Exception:
        pass

    data = await state.get_data()
    institute_id = data.get("institute_id")
    fuclty_id = data.get("fuclty_id")
    study_group_id = data.get("study_group_id")
    education_str = data.get("education")

    institute_name = bd.get_institute_from_id(institute_id) if institute_id else "не выбран"
    fuclty_name_str = bd.get_fuclty_from_id(fuclty_id) if fuclty_id else "не выбрано"
    group_name_str = bd.get_study_group_from_id(study_group_id) if study_group_id else "не выбрана"
    student_name_str = bd.get_student_from_id(student_id) if student_id else "не выбран(а)"
    education_display = education_str.replace('бакалавр', 'Бакалавриат').replace('магистр',
                                                                                 'Магистратура') if education_str else "не выбрана"

    information = (
        f"🎉 <b>Авторизация успешно завершена!</b> 🎉\n\n"
        f"<b>Форма обучения:</b> {education_display}\n"
        f"<b>Институт:</b> {institute_name}\n"
        f"<b>Направление:</b> {fuclty_name_str}\n"
        f"<b>Группа:</b> {group_name_str}\n"
        f"<b>ФИО:</b> {student_name_str}\n\n"
        f"Теперь вы можете пользоваться всеми функциями бота!"
    )
    await callback.message.answer(information, parse_mode="html")
    await callback.message.answer(info_text, reply_markup=kb.first_kb, parse_mode="html")
    await callback.answer("Авторизация завершена!")


# 5. ХЕНДЛЕРЫ ДЛЯ FSM ДОБАВЛЕНИЯ ДАННЫХ (AddData)

@router.callback_query(StateFilter(Student.institute), F.data.startswith("request_add_institute_edu_"))
async def request_add_institute_handler(callback: CallbackQuery, state: FSMContext):
    education_for_add = callback.data.split("request_add_institute_edu_")[1]

    student_data_for_context = await state.get_data()
    await state.set_state(AddData.waiting_for_institute_name)
    await state.update_data(
        add_data_target='institute',
        add_data_education=education_for_add,
        # Сохраняем контекст из Student FSM
        education=student_data_for_context.get("education"),
    )
    await callback.message.edit_text(
        f"Добавление нового института для формы обучения: <b>{education_for_add.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}</b>.\n"
        f"Введите точное название нового института:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✖️ Отмена добавления", callback_data="cancel_add_data_process")]
        ])
    )
    await callback.answer()


@router.message(AddData.waiting_for_institute_name, F.text)
async def process_new_institute_name(message: Message, state: FSMContext, bot: Bot):
    institute_name_input = message.text.strip()
    data = await state.get_data()
    education_for_add = data.get("add_data_education")

    if not institute_name_input or not education_for_add:
        await message.answer("Ошибка: Не удалось получить все данные для заявки. Попробуйте снова.",
                             reply_markup=kb.first_kb)
        await state.clear()
        return

    admin_message_text = (
        f"⚠️ <b>Заявка на добавление ИНСТИТУТА</b> ⚠️\n\n"
        f"<b>Название:</b> {institute_name_input}\n"
        f"<b>Форма обучения:</b> {education_for_add.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"<b>От пользователя:</b> @{message.from_user.username} (ID: {message.from_user.id})"
    )
    try:
        await bot.send_message(ADMIN_USER_ID, admin_message_text, parse_mode="HTML")
        await message.answer(f"Заявка на добавление института '{institute_name_input}' отправлена администратору.",
                             reply_markup=kb.first_kb)
    except Exception as e:
        logging.error(f"Ошибка отправки сообщения админу (институт): {e}")
        await message.answer("Не удалось отправить заявку администратору. Попробуйте позже.", reply_markup=kb.first_kb)

    original_education = data.get("education")  # Восстанавливаем education из Student FSM
    await state.clear()  # Очищаем AddData
    if original_education:  # Если удалось восстановить, возвращаем к выбору института
        await state.set_state(Student.institute)
        await state.update_data(education=original_education)
        await message.answer(
            f"Форма обучения: {original_education.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
            f"2. Выберите ваш институт (список мог обновиться):",
            reply_markup=kb.get_institute_kb(original_education, 1)
        )
    else:  # Иначе в начало
        await state.set_state(Student.education)
        await message.answer("1. Выберите форму обучения:", reply_markup=kb.get_education_kb())


@router.callback_query(StateFilter(Student.fuclty), F.data.startswith("request_add_faculty_inst_"))
async def request_add_faculty_handler(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    institute_id_for_add = int(parts[3])
    education_for_add = parts[5]

    student_data_for_context = await state.get_data()
    await state.set_state(AddData.waiting_for_fuclty_name)
    await state.update_data(
        add_data_target='faculty',
        add_data_institute_id=institute_id_for_add,
        add_data_education=education_for_add,
        education=student_data_for_context.get("education"),
        institute_id=student_data_for_context.get("institute_id")
    )
    institute_name = bd.get_institute_from_id(institute_id_for_add)
    await callback.message.edit_text(
        f"Добавление нового направления для:\n"
        f"<b>Институт:</b> {institute_name}\n"
        f"<b>Форма обучения:</b> {education_for_add.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"Введите точное название нового направления/факультета:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✖️ Отмена добавления", callback_data="cancel_add_data_process")]
        ])
    )
    await callback.answer()


@router.message(AddData.waiting_for_fuclty_name, F.text)
async def process_new_faculty_name(message: Message, state: FSMContext, bot: Bot):
    faculty_name_input = message.text.strip()
    data = await state.get_data()
    institute_id_for_add = data.get("add_data_institute_id")
    education_for_add = data.get("add_data_education")

    if not all([faculty_name_input, institute_id_for_add, education_for_add]):
        await message.answer("Ошибка: Не все данные для заявки (факультет). Попробуйте снова.",
                             reply_markup=kb.first_kb)
        await state.clear()
        return

    institute_name = bd.get_institute_from_id(institute_id_for_add)
    admin_message_text = (
        f"⚠️ <b>Заявка на добавление НАПРАВЛЕНИЯ</b> ⚠️\n\n"
        f"<b>Название:</b> {faculty_name_input}\n"
        f"<b>Институт:</b> {institute_name} (ID: {institute_id_for_add})\n"
        f"<b>Форма обучения:</b> {education_for_add.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"<b>От пользователя:</b> @{message.from_user.username} (ID: {message.from_user.id})"
    )
    try:
        await bot.send_message(ADMIN_USER_ID, admin_message_text, parse_mode="HTML")
        await message.answer(f"Заявка на добавление направления '{faculty_name_input}' отправлена.",
                             reply_markup=kb.first_kb)
    except Exception as e:
        logging.error(f"Ошибка отправки сообщения админу (факультет): {e}")
        await message.answer("Не удалось отправить заявку администратору. Попробуйте позже.", reply_markup=kb.first_kb)

    original_education = data.get("education")
    original_institute_id = data.get("institute_id")
    await state.clear()
    if original_education and original_institute_id:
        await state.set_state(Student.fuclty)
        await state.update_data(education=original_education, institute_id=original_institute_id)
        institute_name_restored = bd.get_institute_from_id(original_institute_id)
        await message.answer(
            f"Форма: {original_education.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
            f"Институт: {institute_name_restored}\n"
            f"3. Выберите ваше направление (список мог обновиться):",
            reply_markup=kb.get_fuclty_kb(original_institute_id, original_education, 1)
        )
    else:
        await state.set_state(Student.education)
        await message.answer("1. Выберите форму обучения:", reply_markup=kb.get_education_kb())


@router.callback_query(StateFilter(Student.group), F.data.startswith("request_add_group_faculty_"))
async def request_add_group_handler(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    faculty_id_for_add = int(parts[4])
    course_for_add = int(parts[6])
    education_for_add = parts[8]

    student_data_for_context = await state.get_data()
    await state.set_state(AddData.waiting_for_group_name)
    await state.update_data(
        add_data_target='group',
        add_data_faculty_id=faculty_id_for_add,
        add_data_course=course_for_add,
        add_data_education=education_for_add,
        education=student_data_for_context.get("education"),
        institute_id=student_data_for_context.get("institute_id"),
        fuclty_id=student_data_for_context.get("fuclty_id"),
        course_for_group_selection=student_data_for_context.get("course_for_group_selection")
    )
    faculty_name = bd.get_fuclty_from_id(faculty_id_for_add)
    institute_name = bd.get_institute_from_id(student_data_for_context.get("institute_id"))
    await callback.message.edit_text(
        f"Добавление новой группы для:\n"
        f"<b>Институт:</b> {institute_name}\n"
        f"<b>Направление:</b> {faculty_name}\n"
        f"<b>Курс:</b> {course_for_add}\n"
        f"<b>Форма:</b> {education_for_add.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"Введите точное название новой группы (например, БИВТ23-1):",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✖️ Отмена добавления", callback_data="cancel_add_data_process")]
        ])
    )
    await callback.answer()


@router.message(AddData.waiting_for_group_name, F.text)
async def process_new_group_name(message: Message, state: FSMContext):
    group_name_input = message.text.strip()
    if not group_name_input:
        await message.answer("Название группы не может быть пустым. Введите снова или отмените.")
        return
    await state.update_data(add_data_group_name=group_name_input)
    await state.set_state(AddData.waiting_for_group_url)
    await message.answer(
        f"Название группы: <b>{group_name_input}</b>.\n"
        f"Теперь введите URL расписания для этой группы (если есть, иначе '-' или 'нет'):",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✖️ Отмена добавления", callback_data="cancel_add_data_process")]
        ])
    )


@router.message(AddData.waiting_for_group_url, F.text)
async def process_new_group_url(message: Message, state: FSMContext, bot: Bot):
    group_url_input = message.text.strip()
    group_url_to_save = group_url_input if group_url_input.lower() not in ['-', 'нет'] else None

    data = await state.get_data()
    group_name = data.get("add_data_group_name")
    faculty_id = data.get("add_data_faculty_id")
    course = data.get("add_data_course")
    education = data.get("add_data_education")
    institute_id = data.get("institute_id")  # Берем из сохраненного Student FSM контекста

    if not all([group_name, faculty_id, course, education, institute_id]):
        await message.answer("Ошибка: Не все данные для заявки (группа). Попробуйте снова.", reply_markup=kb.first_kb)
        await state.clear()
        return

    faculty_name = bd.get_fuclty_from_id(faculty_id)
    institute_name = bd.get_institute_from_id(institute_id)

    admin_message_text = (
        f"⚠️ <b>Заявка на добавление ГРУППЫ</b> ⚠️\n\n"
        f"<b>Название:</b> {group_name}\n"
        f"<b>URL:</b> {group_url_to_save if group_url_to_save else 'Не указан'}\n"
        f"<b>Институт:</b> {institute_name} (ID: {institute_id})\n"
        f"<b>Направление:</b> {faculty_name} (ID: {faculty_id})\n"
        f"<b>Курс:</b> {course}\n"
        f"<b>Форма обучения:</b> {education.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        f"<b>От пользователя:</b> @{message.from_user.username} (ID: {message.from_user.id})"
    )
    try:
        await bot.send_message(ADMIN_USER_ID, admin_message_text, parse_mode="HTML")
        await message.answer(f"Заявка на добавление группы '{group_name}' отправлена.", reply_markup=kb.first_kb)
    except Exception as e:
        logging.error(f"Ошибка отправки сообщения админу (группа): {e}")
        await message.answer("Не удалось отправить заявку администратору. Попробуйте позже.", reply_markup=kb.first_kb)

    original_education = data.get("education")
    original_institute_id = data.get("institute_id")
    original_faculty_id = data.get("fuclty_id")
    original_course = data.get("course_for_group_selection", 1)
    await state.clear()
    if all([original_education, original_institute_id, original_faculty_id]):
        await state.set_state(Student.group)
        await state.update_data(education=original_education, institute_id=original_institute_id,
                                fuclty_id=original_faculty_id, course_for_group_selection=original_course)
        institute_name_restored = bd.get_institute_from_id(original_institute_id)
        faculty_name_restored = bd.get_fuclty_from_id(original_faculty_id)
        await message.answer(
            f"Форма: {original_education.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
            f"Институт: {institute_name_restored}\n"
            f"Направление: {faculty_name_restored}\n"
            f"4. Выберите вашу группу ({original_course} курс, список мог обновиться):",
            reply_markup=kb.get_group_kb(original_faculty_id, original_course, original_education, 1)
        )
    else:
        await state.set_state(Student.education)
        await message.answer("1. Выберите форму обучения:", reply_markup=kb.get_education_kb())


@router.callback_query(StateFilter(Student.name), F.data.startswith("request_add_student_to_group_"))
async def request_add_student_handler(callback: CallbackQuery, state: FSMContext):
    group_id_for_add = int(callback.data.split("_")[-1])
    student_data_for_context = await state.get_data()
    await state.set_state(AddData.waiting_for_student_fio)
    await state.update_data(
        add_data_target='student',
        add_data_group_id=group_id_for_add,
        education=student_data_for_context.get("education"),
        institute_id=student_data_for_context.get("institute_id"),
        fuclty_id=student_data_for_context.get("fuclty_id"),
        study_group_id=student_data_for_context.get("study_group_id"),  # study_group_id это и есть group_id_for_add
        course_for_group_selection=student_data_for_context.get("course_for_group_selection")
    )
    group_name = bd.get_study_group_from_id(group_id_for_add)
    await callback.message.edit_text(
        f"Добавление студента в группу: <b>{group_name}</b>.\n"
        f"Введите ваше полное ФИО (Иванов Иван Иванович):",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✖️ Отмена добавления", callback_data="cancel_add_data_process")]
        ])
    )
    await callback.answer()


@router.message(AddData.waiting_for_student_fio, F.text)
async def process_new_student_fio(message: Message, state: FSMContext,
                                  bot: Bot):  # bot: Bot может быть не нужен, если не отправляем админу
    student_fio_input = message.text.strip()
    data = await state.get_data()
    group_id_for_add = data.get("add_data_group_id")

    education_str = data.get("education")  # Из Student FSM
    institute_id = data.get("institute_id")  # Из Student FSM
    faculty_id = data.get("fuclty_id")  # Из Student FSM
    # Курс нужно определить. Он может быть связан с группой или выбирался ранее.
    # Предположим, что курс группы можно получить из данных группы или он был сохранен в state
    # как course_for_group_selection при выборе группы.
    course_for_db = data.get("course_for_group_selection")  # Это курс, который был активен при выборе группы

    if not all([student_fio_input, group_id_for_add, education_str, institute_id, faculty_id]):
        await message.answer(
            "Ошибка: Не все данные для добавления студента доступны. Попробуйте авторизоваться заново.",
            reply_markup=kb.first_kb)
        await state.clear()
        return

    if course_for_db is None:  # Если курс не был сохранен, это проблема
        logging.warning(
            f"Отсутствует информация о курсе для добавления студента {student_fio_input} в группу {group_id_for_add}")
        # Можно попытаться получить курс из БД по группе, если он там есть,
        # или запросить у пользователя (но это усложнит FSM).
        # Пока что выведем ошибку.
        await message.answer(
            "Ошибка: Не удалось определить курс для добавления студента. Пожалуйста, сообщите администратору.",
            reply_markup=kb.first_kb)
        await state.clear()
        return

    education_id = bd.get_education_id_by_name(education_str)
    if education_id is None:
        await message.answer(
            f"Ошибка: Не удалось определить ID для формы обучения '{education_str}'. Сообщите администратору.",
            reply_markup=kb.first_kb)
        await state.clear()
        return

    # Попытка добавить студента в БД
    new_student_id_from_db = bd.add_new_student_to_db(
        student_fio=student_fio_input,
        id_study_group=group_id_for_add,
        id_fuclty=faculty_id,
        id_institute=institute_id,
        id_education=education_id,
        course=int(course_for_db)  # Убедимся, что курс это число
    )

    if new_student_id_from_db:
        await message.answer(
            f"Вы ('{student_fio_input}') были успешно добавлены в базу данных!\n"
            f"Теперь вы можете авторизоваться, выбрав себя из списка.",
            reply_markup=kb.first_kb
        )
        # Опционально: можно автоматически "авторизовать" пользователя,
        # обновив Student FSM и отправив ему сообщение об успешной авторизации.
        # await state.set_state(None) # Сбрасываем AddData
        # await state.update_data( # Обновляем данные Student FSM
        #     student_id=new_student_id_from_db,
        #     study_group_id=group_id_for_add,
        #     # education, institute_id, fuclty_id уже должны быть в state
        # )
        # await process_student_choice_after_add(message, state) # Специальная функция для сообщения после авто-добавления
        # Но для простоты пока просто сообщаем об успехе.
    else:
        await message.answer(
            f"Не удалось добавить вас ('{student_fio_input}') в базу. "
            f"Возможно, такой студент уже существует в этой группе, или произошла ошибка сервера. "
            f"Пожалуйста, попробуйте выбрать себя из списка или обратитесь к администратору.",
            reply_markup=kb.first_kb
        )

    # Независимо от успеха добавления в БД, очищаем состояние AddData
    # и возвращаем пользователя к предыдущему шагу (выбору студента),
    # чтобы он мог увидеть обновленный список (если добавление было успешно)
    # или попробовать снова / выбрать существующего.

    original_education = data.get("education")
    original_institute_id = data.get("institute_id")
    original_faculty_id = data.get("fuclty_id")
    original_study_group_id = data.get("study_group_id")  # Это ID группы, в которой был список студентов

    await state.clear()  # Очищаем AddData

    if all([original_education, original_institute_id, original_faculty_id, original_study_group_id]):
        await state.set_state(Student.name)  # Возвращаем к выбору студента
        await state.update_data(
            education=original_education,
            institute_id=original_institute_id,
            fuclty_id=original_faculty_id,
            study_group_id=original_study_group_id,
            course_for_group_selection=data.get("course_for_group_selection")  # Восстанавливаем курс
        )
        institute_name_restored = bd.get_institute_from_id(original_institute_id)
        faculty_name_restored = bd.get_fuclty_from_id(original_faculty_id)
        group_name_restored = bd.get_study_group_from_id(original_study_group_id)
        await message.answer(  # Отправляем новое сообщение, т.к. предыдущее было от пользователя
            f"Форма: {original_education.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
            f"Институт: {institute_name_restored}\n"
            f"Направление: {faculty_name_restored}\n"
            f"Группа: {group_name_restored}\n"
            f"5. Выберите себя из списка (список мог обновиться):",
            reply_markup=kb.get_student_kb(original_study_group_id, 1)
        )
    else:  # Если не удалось восстановить контекст, отправляем в начало авторизации
        await state.set_state(Student.education)
        await message.answer("1. Выберите форму обучения:", reply_markup=kb.get_education_kb())

@router.callback_query(F.data == "cancel_add_data_process", StateFilter(AddData))
async def cancel_add_data_fsm_handler(callback: CallbackQuery, state: FSMContext):
    data_before_add = await state.get_data()
    await state.clear()

    restored_education = data_before_add.get("education")
    restored_institute_id = data_before_add.get("institute_id")
    restored_fuclty_id = data_before_add.get("fuclty_id")
    restored_study_group_id = data_before_add.get("study_group_id")
    restored_course = data_before_add.get("course_for_group_selection", 1)
    add_data_target = data_before_add.get("add_data_target")

    message_text = "Добавление отменено."
    reply_mk = None
    target_state_student_fsm = None
    current_message_context = ""  # Для формирования текста сообщения

    # Логика возврата к предыдущему шагу авторизации
    if add_data_target == 'student' and restored_study_group_id and restored_fuclty_id and restored_institute_id and restored_education:
        target_state_student_fsm = Student.name
        # Получаем имена для контекста
        # institute_name = bd.get_institute_from_id(restored_institute_id) # Не нужно в тексте этого сообщения
        # faculty_name = bd.get_fuclty_from_id(restored_fuclty_id)
        # group_name = bd.get_study_group_from_id(restored_study_group_id)
        # current_message_context = f"Институт: {institute_name}\nНаправление: {faculty_name}\nГруппа: {group_name}\n"
        message_text = f"{message_text} Вернулись к выбору студента:"
        reply_mk = kb.get_student_kb(restored_study_group_id, 1)
    elif add_data_target == 'group' and restored_fuclty_id and restored_institute_id and restored_education:
        target_state_student_fsm = Student.group
        institute_name = bd.get_institute_from_id(restored_institute_id)
        faculty_name = bd.get_fuclty_from_id(restored_fuclty_id)
        current_message_context = f"Институт: {institute_name}\nНаправление: {faculty_name}\n"
        message_text = f"{current_message_context}{message_text} Вернулись к выбору группы ({restored_course} курс):"
        reply_mk = kb.get_group_kb(restored_fuclty_id, restored_course, restored_education, 1)
    elif add_data_target == 'faculty' and restored_institute_id and restored_education:
        target_state_student_fsm = Student.fuclty
        institute_name = bd.get_institute_from_id(restored_institute_id)
        current_message_context = f"Институт: {institute_name}\n"
        message_text = f"{current_message_context}{message_text} Вернулись к выбору направления:"
        reply_mk = kb.get_fuclty_kb(restored_institute_id, restored_education, 1)
    elif add_data_target == 'institute' and restored_education:
        target_state_student_fsm = Student.institute
        current_message_context = f"Форма обучения: {restored_education.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
        message_text = f"{current_message_context}{message_text} Вернулись к выбору института:"
        reply_mk = kb.get_institute_kb(restored_education, 1)
    else:
        target_state_student_fsm = Student.education
        message_text = "Добавление данных отменено. Начните авторизацию заново:"
        reply_mk = kb.get_education_kb()

    if target_state_student_fsm:
        await state.set_state(target_state_student_fsm)
        await state.update_data(  # Восстанавливаем все известные данные для Student FSM
            education=restored_education, institute_id=restored_institute_id,
            fuclty_id=restored_fuclty_id, study_group_id=restored_study_group_id,
            course_for_group_selection=restored_course
        )
    try:
        await callback.message.edit_text(message_text, reply_markup=reply_mk, parse_mode="HTML")
    except Exception:
        await callback.message.answer(message_text, reply_markup=reply_mk, parse_mode="HTML")
    await callback.answer("Добавление отменено")


@router.callback_query(StateFilter(Student.group), F.data.startswith("request_add_group_faculty_"))
async def request_add_group_handler(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split("_")
    faculty_id_for_add = int(parts[4])
    course_for_add = int(parts[6])
    education_for_add = parts[8]  # education здесь строка типа "бакалавр"

    student_fsm_data = await state.get_data()
    # Сохраняем текущий контекст Student FSM, чтобы вернуться к нему
    # и передать необходимые ID для создания группы
    await state.set_state(AddData.waiting_for_group_name)  # Переходим в состояние ввода имени группы
    await state.update_data(
        add_data_target='group',
        add_data_faculty_id=faculty_id_for_add,
        add_data_course=course_for_add,
        add_data_education=education_for_add,  # education из callback
        # Сохраняем данные из Student FSM для контекста и возврата
        education=student_fsm_data.get("education"),
        institute_id=student_fsm_data.get("institute_id"),
        fuclty_id=student_fsm_data.get("fuclty_id"),  # Это ID факультета, на котором мы были
        course_for_group_selection=student_fsm_data.get("course_for_group_selection")  # Это курс, который отображался
    )

    faculty_name = bd.get_fuclty_from_id(faculty_id_for_add)
    institute_name = bd.get_institute_from_id(student_fsm_data.get("institute_id"))

    await callback.message.edit_text(
        f"Добавление новой группы для:\n"
        f"<b>Институт:</b> {institute_name if institute_name else 'Не определен'}\n"
        f"<b>Направление:</b> {faculty_name if faculty_name else 'Не определено'}\n"
        f"<b>Курс:</b> {course_for_add}\n"
        f"<b>Форма:</b> {education_for_add.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n\n"
        f"Введите точное название новой группы (например, БИВТ23-1):",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✖️ Отмена добавления", callback_data="cancel_add_data_process")]
        ])
    )
    await callback.answer()


@router.message(AddData.waiting_for_group_name, F.text)
async def process_new_group_name(message: Message, state: FSMContext):
    group_name_input = message.text.strip()
    if not group_name_input:
        await message.answer("Название группы не может быть пустым. Введите снова или отмените.")
        return

    await state.update_data(add_data_group_name_input=group_name_input)  # Используем другое имя, чтобы не путать
    await state.set_state(AddData.waiting_for_group_url)
    await message.answer(
        f"Название группы: <b>{group_name_input}</b>.\n"
        f"Теперь введите URL расписания для этой группы (если есть, иначе '-' или 'нет'):",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="✖️ Отмена добавления", callback_data="cancel_add_data_process")]
        ])
    )


@router.message(AddData.waiting_for_group_url, F.text)
async def process_new_group_url(message: Message, state: FSMContext,
                                bot: Bot):  # bot может быть не нужен, если не шлем админу
    group_url_input = message.text.strip()
    group_url_to_save = group_url_input if group_url_input.lower() not in ['-', 'нет'] else None

    data = await state.get_data()
    new_group_name = data.get("add_data_group_name_input")
    # Данные для контекста и возврата из Student FSM
    faculty_id_for_context = data.get("fuclty_id")  # Это ID факультета, на котором мы были
    course_for_context = data.get("course_for_group_selection")  # Это курс, который отображался
    education_for_context = data.get("education")  # Это форма обучения из Student FSM
    # institute_id_for_context = data.get("institute_id") # Для информации

    if not new_group_name:  # faculty_id, course, education для новой группы берутся из add_data_...
        await message.answer("Ошибка: Не удалось получить название группы. Попробуйте снова.", reply_markup=kb.first_kb)
        await state.clear()
        return

    # Добавляем группу в БД
    new_group_id_from_db = bd.add_new_group_to_db(
        group_name=new_group_name,
        group_url=group_url_to_save
    )

    if new_group_id_from_db:
        await message.answer(
            f"Группа '{new_group_name}' успешно добавлена/найдена в базе данных (ID: {new_group_id_from_db}).\n"
            f"Теперь вы можете выбрать ее из списка или добавить в нее студентов.",
            reply_markup=kb.first_kb  # Или другая клавиатура
        )
        # Опционально: уведомить администратора
        # admin_message_text = f"Добавлена новая группа: {new_group_name} (URL: {group_url_to_save if group_url_to_save else 'не указан'}) пользователем @{message.from_user.username}."
        # try:
        #     await bot.send_message(ADMIN_USER_ID, admin_message_text)
        # except Exception as e:
        #     logging.error(f"Не удалось уведомить админа о добавлении группы: {e}")
    else:
        await message.answer(
            f"Не удалось добавить группу '{new_group_name}' в базу. "
            f"Возможно, произошла ошибка сервера. Попробуйте позже или обратитесь к администратору.",
            reply_markup=kb.first_kb
        )

    # Возвращаем пользователя к выбору группы на том же факультете/курсе
    await state.clear()  # Очищаем AddData состояние
    if faculty_id_for_context and course_for_context and education_for_context:
        await state.set_state(Student.group)
        # Восстанавливаем данные Student FSM
        await state.update_data(
            education=education_for_context,
            institute_id=data.get("institute_id"),  # Восстанавливаем institute_id
            fuclty_id=faculty_id_for_context,
            course_for_group_selection=course_for_context
        )

        institute_name_restored = bd.get_institute_from_id(data.get("institute_id"))
        faculty_name_restored = bd.get_fuclty_from_id(faculty_id_for_context)

        await message.answer(  # Отправляем новое сообщение
            f"Форма: {education_for_context.replace('бакалавр', 'Бакалавриат').replace('магистр', 'Магистратура')}\n"
            f"Институт: {institute_name_restored if institute_name_restored else 'Не определен'}\n"
            f"Направление: {faculty_name_restored if faculty_name_restored else 'Не определено'}\n"
            f"4. Выберите вашу группу ({course_for_context} курс, список мог обновиться):",
            reply_markup=kb.get_group_kb(faculty_id_for_context, course_for_context, education_for_context, 1)
        )
    else:  # Если не удалось восстановить контекст, отправляем в начало авторизации
        await state.set_state(Student.education)
        await message.answer("1. Выберите форму обучения:", reply_markup=kb.get_education_kb())


# 6. ОБЩИЕ ФУНКЦИОНАЛЬНЫЕ ХЕНДЛЕРЫ (Отметиться, Списки, Информация, Расписание)
# Здесь должны быть ваши реализации этих функций, я оставляю заглушки или ваш существующий код
# Важно: убедитесь, что они проверяют, авторизован ли пользователь (т.е. student_id и group_id в state)
chat_ids = [1871402519]


async def clear_schedule(state: FSMContext):
    print("расписание отчистилось")
    await state.update_data(schedule=None)


@router.message(CommandStart())
async def command_start(message: Message, state: FSMContext) -> None:
    """запускается при команде старт"""
    if message.from_user.id not in chat_ids:
        chat_ids.append(message.from_user.id)
    print(chat_ids)
    scheduler = AsyncIOScheduler()
    scheduler.add_job(
        clear_schedule, "cron", minute=1, hour=0, day_of_week="mon", args=[state]
    )
    scheduler.start()
    await message.delete()
    await message.answer(start_text, reply_markup=kb.first_kb)
    await message.answer(info_text, parse_mode="html")
    #await message.answer_sticker(
        #"CAACAgIAAxkBAAEMARFmLVqDWXJu6zTfHeHGH9Ug8Eqx6wACyUAAAr5sEEgMQFhAO8zizTQE"
    #)


# @router.message(Command('help'))
# async def command_help(message: Message) -> None:
#     await message.answer(help_text, parse_mode='HTML')


@router.message(F.text == "Авторизоваться")
async def find_group(message: Message, state: FSMContext) -> None:
    """Хендлер отслеживает слово 'Авторизоваться' и обрабатывает его"""
    await message.delete()
    sent_message = await message.answer(
        text="Начинается авторизация", reply_markup=ReplyKeyboardRemove()
    )
    await sent_message.delete()
    state_data = await state.get_data()
    schedule = state_data.get("schedule")
    if schedule is not None and not (
        schedule == "На эту неделю занятия не поставлены 😘"
    ):
        # await message.answer("")
        print("deleting")
        student_id = state_data.get("student_id")
        group_id = state_data.get("study_group_id")

        for day in schedule:
            times = []
            if not datetime.now().date() == dateparser.parse(day["date"]).date():
                print("Даты не совпадают")
            else:
                for lesson in day["lessons"]:
                    start_h = int(lesson["time_start"].split(":")[0])
                    start_m = int(lesson["time_start"].split(":")[1])
                    end_h = int(lesson["time_end"].split(":")[0])
                    end_m = int(lesson["time_end"].split(":")[1])
                    spot_lesson = lesson["spot"]
                    name_lesson = lesson["name_lesson"]
                    times.append(
                        [
                            time(start_h, start_m),
                            time(end_h, end_m),
                            spot_lesson,
                            name_lesson,
                        ]
                    )
                # times.append([time(1, 00), time(3, 40), "Главный учебный корпус", "Home"])
                # times.append([time(14, 00), time(16, 40), "Главный учебный корпус", "Discord"])

                # print(times)
                for start, end, spot, name in times:
                    print(f"{time(start.hour - 1, start.minute + 40)} - {end} - {spot}")
                    start_time = time(start.hour - 1, start.minute + 40)
                    now_time = time(datetime.now().hour, datetime.now().minute)
                    if start_time <= now_time <= end:
                        bd.delete_visiting(
                            lesson=name,
                            group_id=group_id,
                            student_id=student_id,
                            year=datetime.now().year,
                            month=datetime.now().month,
                            day=datetime.now().day,
                        )

    await state.update_data(
        institute_id=None,
        fuclty_id=None,
        study_group_id=None,
        student_id=None,
        education=None,
        schedule=None,
    )
    await message.answer("Выбери сециальность", reply_markup=kb.get_education_kb())
    # await message.delete()


@router.message(F.text == "Отметиться")
async def get_location(message: Message, state: FSMContext):
    state_data = await state.get_data()
    group_id = state_data.get("study_group_id")
    url = bd.get_schedule_url(group_id)
    if group_id is None:
        await message.answer("Так ты сначала авторизуйся 😡")
        #await message.answer_sticker(
           # "CAACAgQAAxkBAAEMC1dmNK3bWk21XDnN-lEU2XG5EtbuSwACjwcAAvnhgFEO9a_zf5fHtDQE"
        #)
    elif url is None:
        await message.answer("На эту неделю занятия не поставлены 😘")
    else:
        sent_message = await message.answer("Подождите, данные загружаются...")
        schedule_state = state_data.get("schedule")
        if schedule_state is None:
            schedule = purse.get_schedule(url)
            await state.update_data(schedule=schedule)
        state_data = await state.get_data()
        schedule = state_data.get("schedule")
        await sent_message.delete()

        if schedule == "На эту неделю занятия не поставлены 😘":
            await message.answer(schedule)
        else:
            flag = True
            for day in schedule:
                times = []
                if not datetime.now().date() == dateparser.parse(day["date"]).date():
                    print("Даты не совпадают")
                else:
                    print("Даты совпадают")
                    for lesson in day["lessons"]:

                        start_h = int(lesson["time_start"].split(":")[0])
                        start_m = int(lesson["time_start"].split(":")[1])
                        end_h = int(lesson["time_end"].split(":")[0])
                        end_m = int(lesson["time_end"].split(":")[1])
                        spot_lesson = lesson["spot"]
                        name_lesson = lesson["name_lesson"]

                        times.append(
                            [
                                time(start_h, start_m),
                                time(end_h, end_m),
                                spot_lesson,
                                name_lesson,
                            ]
                        )

                    # times.append([time(14, 00), time(16, 40), "Главный учебный корпус", "Discord"])

                    for start, end, spot, name in times:
                        # print(f"{time(start.hour - 1, start.minute + 40)} - {end} - {spot}")
                        start_time = time(start.hour - 1, start.minute + 40)
                        now_time = time(datetime.now().hour, datetime.now().minute)
                        if start_time <= now_time <= end:
                            flag = False
                            # print(f"now: {time(start.hour - 1, start.minute + 40)} - {end} - {spot}")
                            mess_text = (
                                "Вы собираетесь отметиться на паре \n\n<b>Предмет:</b> "
                                + name
                                + "\n\n<b>Время:</b> "
                                + start.strftime("%H:%M")
                                + "-"
                                + end.strftime("%H:%M")
                                + "\n\n<b>Место проведения:</b> "
                                + spot
                            )
                            await message.answer(mess_text, parse_mode="html")
                            await message.answer(
                                "Пожалуйста, отправьте ваше местоположение.",
                                reply_markup=kb.geo_kb,
                            )

            if flag:
                await message.answer(
                    "В данный момент ты ни на что не можешь отметиться",
                    reply_markup=kb.cancel_kb,
                )
                #sticker_id = "CAACAgIAAxkBAAEMETpmOojZfGmJtZsVqrGTqV1AB1tangACORYAAjL-yEpvoPRnuqpSqTUE"
                #await message.answer_sticker(sticker_id)


@router.message(F.location)
async def handle_location(message: Message, state: FSMContext):
    latitude = message.location.latitude
    longitude = message.location.longitude
    state_data = await state.get_data()
    schedule = state_data.get("schedule")
    student_id = state_data.get("student_id")
    group_id = state_data.get("study_group_id")
    if not schedule == "На эту неделю занятия не поставлены 😘":
        for day in schedule:
            times = []
            if not datetime.now().date() == dateparser.parse(day["date"]).date():
                print("Даты не совпадают")
            else:
                for lesson in day["lessons"]:
                    start_h = int(lesson["time_start"].split(":")[0])
                    start_m = int(lesson["time_start"].split(":")[1])
                    end_h = int(lesson["time_end"].split(":")[0])
                    end_m = int(lesson["time_end"].split(":")[1])
                    spot_lesson = lesson["spot"]
                    name_lesson = lesson["name_lesson"]
                    times.append(
                        [
                            time(start_h, start_m),
                            time(end_h, end_m),
                            spot_lesson,
                            name_lesson,
                        ]
                    )
                # times.append([time(1, 00), time(3, 40), "Главный учебный корпус", "Home"])
                # times.append([time(14, 00), time(16, 40), "Главный учебный корпус", "Discord"])

                # print(times)
                for start, end, spot, name in times:
                    print(f"{time(start.hour - 1, start.minute + 40)} - {end} - {spot}")
                    start_time = time(start.hour - 1, start.minute + 40)
                    now_time = time(datetime.now().hour, datetime.now().minute)
                    if start_time <= now_time <= end:
                        geo = bd.get_geo(spot)
                        spot_geo = (float(geo[0]), float(geo[1]))
                        student_geo = (latitude, longitude)
                        distance = haversine(spot_geo, student_geo, unit=Unit.METERS)
                        print(distance)

                        if distance > 350:
                            mess = (
                                "Ты НЕ смог отметиться на паре\n\n"
                                + "Чтобы отметиться нужно находиться до 350 метров от корпуса\n\n"
                                + "Твое расстояние до  учебного корпуса: ~"
                                + str(round(distance))
                                + " метров"
                            )
                            await message.answer(mess, reply_markup=kb.geo_kb)
                        else:
                            success = bd.insert_lesson(
                                lesson_name=name,  # Изменено с 'lesson' на 'lesson_name'
                                group_id=group_id,
                                student_id=student_id,
                                year=datetime.now().year,
                                month=datetime.now().month,
                                day=datetime.now().day,
                            )
                            if success:
                                await message.answer(f"Вы успешно отмечены на паре: {name}!", reply_markup=kb.first_kb)
                            else:
                                await message.answer(
                                    f"Не удалось отметиться на паре: {name}. Возможно, вы уже отмечены или произошла ошибка.",
                                    reply_markup=kb.first_kb)


@router.message(F.text == "Списки присутствующих")
async def send_file_handler(message: Message, state: FSMContext) -> None:
    state_data = await state.get_data()
    group_id = state_data.get("study_group_id")

    if group_id is None:
        await message.answer("Так ты сначала авторизуйся 😡 (нужно выбрать ФИО)", reply_markup=kb.first_kb)
        return

    sent_progress_message = None
    try:
        sent_progress_message = await message.answer(
            "Формирую файл со списками присутствующих... Пожалуйста, подождите."
        )
    except Exception as e:
        print(f"Не удалось отправить сообщение о прогрессе: {e}")

    book = openpyxl.Workbook()
    if "Sheet" in book.sheetnames:
        try:
            book.remove(book["Sheet"])
        except Exception as e_remove_sheet:
            print(f"Не удалось удалить стандартный лист 'Sheet': {e_remove_sheet}")
            # Если удалить не удалось, просто продолжаем, возможно, он будет перезаписан или останется пустым

    lessons = bd.get_lessons(group_id)
    print(f"[send_file_handler] Для группы ID {group_id} найдены предметы: {lessons}")

    data_found_for_any_sheet = False

    if lessons:
        for lesson_name_str in lessons:
            safe_sheet_name = "".join(c if c.isalnum() or c in [' ', '-'] else '_' for c in lesson_name_str)
            safe_sheet_name = safe_sheet_name[:30]
            if not safe_sheet_name:
                safe_sheet_name = f"Предмет_{lessons.index(lesson_name_str) + 1}"

            sheet_title_candidate = safe_sheet_name
            counter = 1
            while sheet_title_candidate in book.sheetnames:
                sheet_title_candidate = f"{safe_sheet_name[:28]}_{counter}"
                counter += 1
            sheet = book.create_sheet(sheet_title_candidate)
            print(f"[send_file_handler] Создан лист: {sheet_title_candidate} для предмета '{lesson_name_str}'")

            current_date_obj = datetime.now().date()
            year_for_filter = current_date_obj.year if current_date_obj.month >= 9 else current_date_obj.year - 1
            date_start_filter = f"{year_for_filter}-09-01"

            print(
                f"  Вызов get_lesson_dates с: lesson_name='{lesson_name_str}', group_id={group_id}, date_start_str='{date_start_filter}'")
            lesson_dates_as_header = bd.get_lesson_dates(
                lesson_name=lesson_name_str,
                group_id=group_id,
                date_start_str=date_start_filter
            )
            print(f"  Получены даты (с заголовком 'Студент') для '{lesson_name_str}': {lesson_dates_as_header}")

            actual_dates_to_query = lesson_dates_as_header[1:] if len(lesson_dates_as_header) > 1 else []

            student_attendance_rows = []
            if actual_dates_to_query:
                print(
                    f"  Вызов get_student_list с: dates_dd_mm_yyyy={actual_dates_to_query}, group_id={group_id}, lesson_name='{lesson_name_str}'")
                student_attendance_rows = bd.get_student_list(
                    dates_dd_mm_yyyy=actual_dates_to_query,
                    group_id=group_id,
                    lesson_name=lesson_name_str
                )
                if student_attendance_rows:
                    data_found_for_any_sheet = True
            print(f"  Данные посещаемости для '{lesson_name_str}': {student_attendance_rows}")

            excel_rows_to_write = [lesson_dates_as_header]
            if student_attendance_rows:
                excel_rows_to_write.extend(student_attendance_rows)

            if len(excel_rows_to_write) == 1 and not actual_dates_to_query:
                sheet.append(["Нет данных о посещаемости для этого предмета."])
            else:
                for row_content in excel_rows_to_write:
                    sheet.append(row_content)

                for col_idx, column_cells in enumerate(sheet.columns):
                    max_cell_length = 0
                    column_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                    for cell in column_cells:
                        try:
                            if cell.value is not None:
                                cell_value_as_str = str(cell.value)
                                if len(cell_value_as_str) > max_cell_length:
                                    max_cell_length = len(cell_value_as_str)
                        except Exception as e_cell_format:
                            print(f"Ошибка при форматировании ячейки {cell.coordinate}: {e_cell_format}")
                    adjusted_column_width = (max_cell_length + 2) * 1.2 if max_cell_length > 0 else 10
                    sheet.column_dimensions[column_letter].width = adjusted_column_width

                for row_object in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                    for cell in row_object:
                        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

                if sheet.max_row >= 1:
                    for cell in sheet[1]:
                        cell.font = openpyxl.styles.Font(bold=True)

    if not data_found_for_any_sheet:
        if not lessons:
            summary_sheet = book.create_sheet("Нет данных") if "Нет данных" not in book.sheetnames else book[
                "Нет данных"]
            summary_sheet.append(["Для вашей группы не найдено предметов в базе данных."])
        else:
            summary_sheet = book.create_sheet(
                "Сводка_нет_посещений") if "Сводка_нет_посещений" not in book.sheetnames else book[
                "Сводка_нет_посещений"]
            summary_sheet.append(["По выбранным предметам не найдено данных о посещаемости за указанный период."])
        print("[send_file_handler] Данных для Excel не найдено или не было уроков с посещениями.")

    excel_buffer = io.BytesIO()
    try:
        book.save(excel_buffer)
        excel_buffer.seek(0)  # Важно! Переместить указатель в начало буфера для чтения
        print(f"[send_file_handler] Excel книга сохранена в буфер. Размер: {excel_buffer.getbuffer().nbytes} байт.")
    except Exception as e_save:
        print(f"[send_file_handler] Ошибка при сохранении Excel в буфер: {e_save}")
        if sent_progress_message:
            try:
                await sent_progress_message.delete()
            except:
                pass
        await message.answer("Произошла ошибка при формировании отчета. Пожалуйста, попробуйте позже.",
                             reply_markup=kb.first_kb)
        return

    if sent_progress_message:
        try:
            await sent_progress_message.delete()
        except:
            pass

    group_name_from_db = bd.get_study_group_from_id(group_id) or "UnknownGroup"
    safe_group_name_for_file = "".join(c if c.isalnum() else "_" for c in group_name_from_db)
    final_file_name = f"Посещаемость_{safe_group_name_for_file}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

    # Порог для проверки, не пустой ли файл. 22 байта - это пустой zip. Реальный xlsx будет больше.
    # Файл с одним листом и одной строкой "Нет данных" весит около 5-7 Кбайт.
    if excel_buffer.getbuffer().nbytes > 500:  # Увеличил порог для большей надежности
        try:
            # Используем BufferedInputFile для отправки из io.BytesIO
            document_to_send = BufferedInputFile(excel_buffer.read(), filename=final_file_name)
            await message.answer_document(
                document=document_to_send,
                caption="Файл со списками присутствующих за текущий учебный год.",
                reply_markup=kb.first_kb
            )
            print(f"[send_file_handler] Файл '{final_file_name}' успешно отправлен из буфера.")
        except Exception as e_send:
            print(f"[send_file_handler] Ошибка при отправке документа: {e_send}")
            await message.answer("Произошла ошибка при отправке файла отчета. Пожалуйста, попробуйте позже.",
                                 reply_markup=kb.first_kb)
    else:
        print(
            f"[send_file_handler] Буфер Excel пуст или содержит слишком мало данных ({excel_buffer.getbuffer().nbytes} байт). Файл не будет отправлен.")
        await message.answer(
            "Не удалось сформировать файл: нет данных для отображения или произошла внутренняя ошибка.",
            reply_markup=kb.first_kb)

    excel_buffer.close()


@router.message(F.text == "Информация")
async def print_info(message: Message, state: FSMContext) -> None:
    # await message.delete()
    state_data = await state.get_data()

    institute = "Неизвестно"
    institute_id = state_data.get("institute_id")
    if institute_id is not None:
        institute = bd.get_institute_from_id(institute_id)

    fuclty = "Неизвестно"
    fuclty_id = state_data.get("fuclty_id")
    if fuclty_id is not None:
        fuclty = bd.get_fuclty_from_id(fuclty_id)

    group = "Неизвестно"
    study_group_id = state_data.get("study_group_id")
    if study_group_id is not None:
        group = bd.get_study_group_from_id(study_group_id)

    student = "Неизвестно"
    student_id = state_data.get("student_id")
    if student_id is not None:
        student = bd.get_student_from_id(student_id)

    education = "Неизвестно"
    education_from_state = state_data.get("education")
    if education_from_state is not None:
        education = education_from_state

    print(f"institute: {institute_id}")
    print(f"fuclty: {fuclty_id}")
    print(f"group: {study_group_id}")
    print(f"student: {student_id}")
    information = (
        """Информация про тебя:\n\n<b>Институт:</b>\n"""
        + institute
        + "\n\n<b>Направление:</b>\n"
        + fuclty
        + "\n\n<b>Группа:</b>\n"
        + group
        + "\n\n<b>Уровень образования:</b>\n"
        + education
        + "\n\nФИО:\n"
        + student
    )
    await message.answer(information, parse_mode="html")
    # await message.delete()


@router.message(F.text == "Расписание")
async def show_schedule(message: Message, state: FSMContext):
    state_data = await state.get_data()
    group_id = state_data.get("study_group_id")
    url = bd.get_schedule_url(group_id)
    print(url)
    if group_id is None:
        await message.answer("Так ты сначала авторизуйся 😡")
        #await message.answer_sticker(
            #"CAACAgQAAxkBAAEMC1dmNK3bWk21XDnN-lEU2XG5EtbuSwACjwcAAvnhgFEO9a_zf5fHtDQE"
        #)
    elif url is None:
        await message.answer("На эту неделю занятия не поставлены 😘")
    else:
        sent_message = await message.answer("Подождите, данные загружаются...")
        schedule_state = state_data.get("schedule")
        if schedule_state is None:
            schedule = purse.get_schedule(url)
            await state.update_data(schedule=schedule)
        state_data = await state.get_data()
        schedule = state_data.get("schedule")
        if schedule == "На эту неделю занятия не поставлены 😘":
            await message.answer(schedule)
        else:
            texts = []
            for day in schedule:
                text = "<b>" + day["date"] + "</b>\n\n\t"
                for lesson in day["lessons"]:
                    text += lesson["time_start"] + "-" + lesson["time_end"] + "\n\t"
                    text += lesson["name_lesson"] + "\n\t"
                    text += lesson["lesson_type"] + "\n\t"
                    if not lesson["teacher"] == "":
                        text += lesson["teacher"] + "\n\t"
                    text += lesson["spot"] + ", ауд. " + lesson["auditory"] + "\n\n"
                texts.append(text)
            await sent_message.delete()
            for elem in texts:
                await message.answer(elem, parse_mode="html")


# ... (Ваш код для Отметиться, Списки, Информация, Расписание) ...
# Пример для Информации, остальное по аналогии

# 7. ОБЩИЙ ХЕНДЛЕР ДЛЯ КНОПКИ "✖️Отмена" (ReplyKeyboard)
@router.message(F.text == "✖️Отмена")
async def handle_reply_cancel_button(message: Message, state: FSMContext):
    current_fsm_state = await state.get_state()
    action_cancelled_message = "Действие отменено."

    if current_fsm_state:
        if current_fsm_state.startswith("Student:"):
            action_cancelled_message = "Процесс авторизации прерван и сброшен."
        elif current_fsm_state.startswith("AddData:"):
            action_cancelled_message = "Процесс добавления данных прерван."
        await state.clear()
        await message.answer(action_cancelled_message, reply_markup=kb.first_kb)
    else:
        await message.answer("Нет активных действий для отмены. Вы в главном меню.", reply_markup=kb.first_kb)
    try:
        await message.delete()
    except Exception:
        pass
    # await message.answer(info_text, reply_markup=kb.first_kb, parse_mode="html") # info_text можно не дублировать


# 8. ОБЩИЙ ХЕНДЛЕР ДЛЯ НЕИЗВЕСТНЫХ КОЛБЭКОВ
@router.callback_query()
async def handle_unknown_callback(callback: CallbackQuery, state: FSMContext):
    current_fsm_state = await state.get_state()
    logging.warning(f"Необработанный callback: {callback.data}, текущее состояние: {current_fsm_state}")
    await callback.answer("Это действие сейчас не поддерживается или кнопка устарела.", show_alert=True)
    if not current_fsm_state:
        try:
            await callback.message.edit_reply_markup(reply_markup=None)
        except:
            pass
        await callback.message.answer("Пожалуйста, используйте основное меню.", reply_markup=kb.first_kb)


# 9. ОБЩИЙ ХЕНДЛЕР ДЛЯ НЕИЗВЕСТНЫХ СООБЩЕНИЙ (echo)
@router.message()
async def echo_message_handler(message: Message, state: FSMContext) -> None:
    current_fsm_state = await state.get_state()
    if current_fsm_state:
        state_name_readable = current_fsm_state.split(':')[-1]
        await message.answer(
            f"Вы находитесь в процессе ({state_name_readable}).\n"
            f"Пожалуйста, завершите его, используя предложенные кнопки, или нажмите '✖️Отмена'."
        )
    else:
        await message.answer(text="Извините, я не понимаю эту команду. Пожалуйста, используйте кнопки.")

